# core/planning/template_builder.py
import pandas as pd
from pathlib import Path
import sys
from core.logger import logger
from openpyxl.styles import NamedStyle, Font, PatternFill  # type: ignore
from openpyxl.utils import get_column_letter  # type: ignore

class TradingPlanTemplate:
    def __init__(self):
        self.template_dir = Path("plans/templates")
        self.template_dir.mkdir(parents=True, exist_ok=True)
        self.default_values = {
            'risk_per_trade': 0.01,
            'profit_to_loss_ratio': 2.0,
            'available_quantity_ratio': 0.8
        }

    def _setup_workbook_styles(self, workbook):
        header_style = NamedStyle(name="header")
        header_style.font = Font(bold=True)
        header_style.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

        pct_style = NamedStyle(name="percentage", number_format='0.00%')
        ratio_style = NamedStyle(name="ratio", number_format='0.00')

        # Avoid duplicate style errors
        for style in [header_style, pct_style, ratio_style]:
            if style.name not in workbook.named_styles:
                workbook.add_named_style(style)

        return header_style, pct_style, ratio_style
    
    def create_template(self, accounts: list):
        """Generate multi-tab Excel template for given accounts with risk parameters"""
        try:
            template_path = self.template_dir / "trading_plan_template.xlsx"

            columns = [
                "Symbol",
                "Entry Price",
                "Stop Loss",
                "Take Profit",
                "Quantity",
                "Risk Per Trade",
                "Profit/Loss Ratio",
                "Available Qty Ratio",
                "Expiry Date",
                "Notes"
            ]
            df = pd.DataFrame(columns=columns)

            with pd.ExcelWriter(template_path, engine='openpyxl', mode='w') as writer:
                # Write each account sheet
                for account in accounts:
                    account_df = df.copy()
                    account_df.to_excel(
                        writer,
                        sheet_name=str(account),
                        index=False,
                        header=True
                    )

                if not writer.sheets:
                    raise ValueError("No worksheets were created in the template. Check account list.")

                workbook = writer.book
                header_style, pct_style, ratio_style = self._setup_workbook_styles(workbook)

                # Apply formatting
                col_settings = [
                    ("Symbol", 15),
                    ("Entry Price", 12, "#,##0.00"),
                    ("Stop Loss", 12, "#,##0.00"),
                    ("Take Profit", 12, "#,##0.00"),
                    ("Quantity", 10, "0"),
                    ("Risk Per Trade", 15, "0.00%"),
                    ("Profit/Loss Ratio", 15, "0.00"),
                    ("Available Qty Ratio", 15, "0.00"),
                    ("Expiry Date", 12, "yyyy-mm-dd"),
                    ("Notes", 30)
                ]

                for sheet_name in writer.sheets:
                    worksheet = writer.sheets[sheet_name]

                    # Header style
                    for col in range(1, len(columns) + 1):
                        worksheet.cell(row=1, column=col).style = header_style

                    # Widths and formats
                    for idx, setting in enumerate(col_settings, start=1):
                        col_letter = get_column_letter(idx)
                        worksheet.column_dimensions[col_letter].width = setting[1]
                        if len(setting) > 2:
                            for row in range(2, 100):
                                cell = worksheet.cell(row=row, column=idx)
                                cell.number_format = setting[2]

            logger.info(f"Template created at {template_path}")
            return template_path

        except ImportError as e:
            logger.critical(f"Required package missing: {str(e)}\nRun: pip install openpyxl pandas")
            sys.exit(1)
        except Exception as e:
            logger.critical(f"Template creation failed: {str(e)}")
            sys.exit(1)

    def create_example_template(self):
        """Create an example template with sample data including risk parameters"""
        example_path = self.template_dir / "example_plan.xlsx"

        example_data = {
            "Symbol": ["AAPL", "MSFT", "TSLA"],
            "Entry Price": [150.0, 300.0, 250.0],
            "Stop Loss": [145.0, 290.0, 240.0],
            "Take Profit": [160.0, 320.0, 270.0],
            "Quantity": [100, 50, 75],
            "Risk Per Trade": [0.015, 0.008, 0.02],
            "Profit/Loss Ratio": [2.0, 2.5, 1.8],
            "Available Qty Ratio": [0.9, 0.75, 0.85],
            "Expiry Date": ["2023-12-15", "2023-12-20", "2023-12-18"],
            "Notes": [
                "Tech breakout play",
                "Support bounce with earnings",
                "High volatility trade"
            ]
        }

        try:
            with pd.ExcelWriter(example_path, engine='openpyxl') as writer:
                df = pd.DataFrame(example_data)
                df.to_excel(writer, index=False, sheet_name="Example Trades")

                workbook = writer.book
                worksheet = writer.sheets["Example Trades"]

                for row in range(2, len(example_data["Symbol"]) + 2):
                    worksheet.cell(row=row, column=6).number_format = '0.00%'

                for col, width in enumerate([15, 12, 12, 12, 10, 15, 15, 15, 12, 30], start=1):
                    worksheet.column_dimensions[get_column_letter(col)].width = width

            logger.info(f"Example template created at {example_path}")
            return example_path

        except Exception as e:
            logger.error(f"Failed to create example template: {str(e)}")
            raise
